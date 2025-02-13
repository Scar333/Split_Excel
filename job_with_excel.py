import os
import openpyxl
from datetime import datetime


def check_folder(path_to_folder: str) -> None:
    """Проверка на существовании папки"""
    if not os.path.exists(path_to_folder):
        os.makedirs(path_to_folder, exist_ok=True)


def auto_size_excel_file(path_to_file: str) -> None:
    """Автоширина Excel файла"""
    wb = openpyxl.load_workbook(path_to_file)
    ws = wb.active
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width
    wb.save(path_to_file)


def format_date(value):
    """Форматирование даты"""
    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y')
    return value


def creat_new_excel_file(key: list, data: list, name_file: str):
    wb = openpyxl.Workbook()
    ws = wb.active

    for col, header in enumerate(key, 1):
        ws.cell(row=1, column=col, value=header)

    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            formatted_value = format_date(value)
            ws.cell(row=row, column=col, value=formatted_value)

    current_date_and_time = datetime.now().strftime('%d_%m_%Y')
    path_to_folder = os.path.join(os.getcwd(), current_date_and_time)
    check_folder(path_to_folder)
    new_path_to_file = os.path.join(path_to_folder, name_file)
    wb.save(new_path_to_file)
    auto_size_excel_file(new_path_to_file)


def creat_new_files(path_to_file: str, count: int):
    name_file = os.path.basename(path_to_file)
    wb = openpyxl.load_workbook(path_to_file)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    df_keys = [cell.value for cell in ws[1]]

    number_two = len(data)

    tmp_result = number_two // count

    tmp_num = 0
    tmp_num_2 = 0
    for i in range(count):
        if i == 0:
            creat_new_excel_file(key=df_keys, data=data[0:tmp_result], name_file=f'{i + 1}_{name_file}')
        elif i == count - 1:
            creat_new_excel_file(key=df_keys, data=data[tmp_num:], name_file=f'{i + 1}_{name_file}')
        else:
            creat_new_excel_file(key=df_keys, data=data[tmp_num:tmp_num_2], name_file=f'{i + 1}_{name_file}')

        tmp_num += tmp_result
        tmp_num_2 = tmp_num + tmp_result
